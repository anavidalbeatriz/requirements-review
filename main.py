import os
import json
import traceback
from dotenv import load_dotenv
from docx import Document
import pandas as pd
from openai import OpenAI
from PyPDF2 import PdfReader
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule

# --- Load API key ---
load_dotenv()
api_key = os.getenv("OPENAI_API_KEY")
if not api_key:
    raise ValueError("OPENAI_API_KEY not found in .env file")

# Initialize OpenAI client
client = OpenAI(api_key=api_key)

# --- Load rules ---
def load_rules(path="rules.json"):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

# --- Read DOCX ---
def read_docx(path):
    doc = Document(path)
    return "\n".join([p.text for p in doc.paragraphs])

# --- Read PDF ---
def read_pdf(path):
    text = ""
    try:
        with open(path, "rb") as f:
            reader = PdfReader(f)
            for page in reader.pages:
                text += page.extract_text() or ""
    except Exception as e:
        print(f"⚠️ Error reading PDF {path}: {e}")
    return text.strip()

# --- Evaluate one rule ---
def evaluate_rule(document_text, rule):
    prompt = f"""
You are an evaluator. Assess the following document based on this rule.

Rule: {rule['name']}
Description: {rule['description']}
Scoring Criteria:
{json.dumps(rule['criteria'], indent=2)}

Document:
{document_text[:8000]}

Respond in strict JSON format:
{{
  "score": 0-3,
  "justification": "Brief explanation of why this score was chosen"
}}
"""
    try:
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "user", "content": prompt}],
            temperature=0
        )
        content = response.choices[0].message.content.strip()
        result = json.loads(content)

        if not isinstance(result.get("score"), int) or "justification" not in result:
            raise ValueError("Invalid JSON structure returned by model")

    except Exception as e:
        print(f"⚠️ Error evaluating rule '{rule['name']}': {e}")
        traceback.print_exc()
        result = {
            "score": None,
            "justification": f"Evaluation failed: {str(e)}"
        }

    return result

# --- Evaluate all rules ---
def evaluate_document(document_text, rules):
    results = {}
    for rule in rules:
        result = evaluate_rule(document_text, rule)
        results[rule["description"]] = result["score"]
    return results

# --- Apply conditional formatting ---
def apply_conditional_formatting(excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active

    # Find the "Total Score" column
    headers = [cell.value for cell in ws[1]]
    if "Total Score" not in headers:
        wb.save(excel_path)
        return

    col_index = headers.index("Total Score") + 1
    col_letter = chr(64 + col_index)
    last_row = ws.max_row

    # Apply green-to-red gradient (low = red, high = green)
    color_scale = ColorScaleRule(
        start_type="min", start_color="F8696B",  # red
        mid_type="percentile", mid_value=50, mid_color="FFEB84",  # yellow
        end_type="max", end_color="63BE7B"  # green
    )

    ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{last_row}", color_scale)
    wb.save(excel_path)
    print("🎨 Conditional formatting applied (red → green gradient on Total Score)")

# --- Main function ---
def main(documents_folder="documents", rules_file="rules.json", output_file="evaluation.xlsx"):
    rules = load_rules(rules_file)
    all_results = []

    for filename in os.listdir(documents_folder):
        if filename.lower().endswith((".docx", ".pdf")):
            file_path = os.path.join(documents_folder, filename)
            print(f"\n📝 Evaluating document: {filename}")

            if filename.lower().endswith(".docx"):
                text = read_docx(file_path)
            elif filename.lower().endswith(".pdf"):
                text = read_pdf(file_path)
            else:
                continue

            evaluation = evaluate_document(text, rules)

            # Build row: document name + rule scores + total
            evaluation_row = {"Document Name": filename}
            evaluation_row.update(evaluation)
            scores = [v for v in evaluation.values() if isinstance(v, (int, float))]
            evaluation_row["Total Score"] = sum(scores) if scores else None

            all_results.append(evaluation_row)

    # Save to Excel
    df = pd.DataFrame(all_results)
    df.to_excel(output_file, index=False)
    print(f"\n✅ Evaluation saved to Excel file: {output_file}")

    # Add conditional formatting
    apply_conditional_formatting(output_file)

if __name__ == "__main__":
    main()
